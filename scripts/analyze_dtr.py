import openpyxl
import json
import os

def analyze_excel(file_path):
    if not os.path.exists(file_path):
        return {"error": f"File {file_path} not found"}

    wb = openpyxl.load_workbook(file_path, data_only=False)
    analysis = {
        "sheets": []
    }

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_data = {
            "name": sheet_name,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "merged_cells": [str(mc) for mc in sheet.merged_cells.ranges],
            "column_dimensions": {},
            "row_dimensions": {},
            "cells": []
        }

        # Get column widths
        for col_letter, col_dim in sheet.column_dimensions.items():
            sheet_data["column_dimensions"][col_letter] = col_dim.width

        # Get row heights
        for row_index, row_dim in sheet.row_dimensions.items():
            sheet_data["row_dimensions"][row_index] = row_dim.height

        # Sample cells for content and styling
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None or cell.has_style:
                    cell_info = {
                        "coordinate": cell.coordinate,
                        "value": str(cell.value) if cell.value is not None else None,
                        "data_type": cell.data_type,
                        "number_format": cell.number_format,
                    }
                    
                    if cell.has_style:
                        style = {
                            "font": {
                                "name": cell.font.name,
                                "size": cell.font.size,
                                "bold": cell.font.bold,
                                "italic": cell.font.italic,
                                "color": str(cell.font.color.rgb) if cell.font.color and hasattr(cell.font.color, 'rgb') else None
                            },
                            "fill": {
                                "fill_type": cell.fill.fill_type,
                                "fgColor": str(cell.fill.fgColor.rgb) if cell.fill.fgColor and hasattr(cell.fill.fgColor, 'rgb') else None
                            },
                            "alignment": {
                                "horizontal": cell.alignment.horizontal,
                                "vertical": cell.alignment.vertical,
                                "wrap_text": cell.alignment.wrap_text
                            },
                            "border": {
                                "left": str(cell.border.left.style) if cell.border.left else None,
                                "right": str(cell.border.right.style) if cell.border.right else None,
                                "top": str(cell.border.top.style) if cell.border.top else None,
                                "bottom": str(cell.border.bottom.style) if cell.border.bottom else None
                            }
                        }
                        cell_info["style"] = style
                    
                    sheet_data["cells"].append(cell_info)
        
        analysis["sheets"].append(sheet_data)

    return analysis

if __name__ == "__main__":
    result = analyze_excel("output/DTR.xlsx")
    print(json.dumps(result, indent=2))
